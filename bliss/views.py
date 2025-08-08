import os
import datetime
from django.shortcuts import render, redirect, get_object_or_404
from .models import Bliss
from .forms import BlissForm
from django.http import HttpResponse
from openpyxl import load_workbook
from django.db.models import Count, Sum
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.db.models import F
from django.db.models import Count, Sum, F, FloatField, DecimalField
from django.db.models.functions import Coalesce
from decimal import Decimal, DivisionByZero, InvalidOperation
from django.contrib import messages
from django.db import models

# Listar
def index(request):
    registros = Bliss.objects.all()
    return render(request, 'bliss/index.html', {'registros': registros})

# Criar
def bliss_create(request):
    if request.method == 'POST':
        form = BlissForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('index')
    else:
        form = BlissForm()
    return render(request, 'bliss/form.html', {'form': form})

# Editar
def bliss_update(request, pk):
    registro = get_object_or_404(Bliss, pk=pk)
    form = BlissForm(request.POST or None, instance=registro)
    if form.is_valid():
        form.save()
        return redirect('index')
    return render(request, 'bliss/form.html', {'form': form})

# Excluir
def bliss_delete(request, pk):
    registro = get_object_or_404(Bliss, pk=pk)
    if request.method == 'POST':
        registro.delete()
        return redirect('index')
    return render(request, 'bliss/delete_confirm.html', {'registro': registro})

# Importar planilha Excel
def bliss_import(request):
    
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        linhas=0
        for row in ws.iter_rows(min_row=2, values_only=True):
            linhas += 1 
            if linhas > 85:
                break
            if not row[0]:  # pula linhas vazias
                continue
            Bliss.objects.create(
                bloco=row[0],
                unidade=row[1],
                area_privativa=row[2],
                garagem=row[3],
                deposito=row[4],
                tipologia=row[5],
                situacao=row[6],
                valor_tabela=row[7] or 0,
                data_venda=row[8] if isinstance(row[8], datetime.date) else None,
                valor_venda=row[9] or 0,
                cliente=row[10] or '',
                email=row[11] or ''
            )
        return redirect('index')
    return render(request, 'bliss/import.html')

# Relatório HTML
def bliss_report(request):
    sort = request.GET.get('sort', 'bloco')
    direction = request.GET.get('dir', 'asc')

    if direction == 'desc':
        sort_order = '-' + sort
    else:
        sort_order = sort

    registros = Bliss.objects.all().order_by(sort_order)

    context = {
        'registros': registros,
        'sort': sort,
        'dir': direction,
        'invert_dir': 'desc' if direction == 'asc' else 'asc',
    }
    return render(request, 'bliss/report.html', context)

# Relatório PDF
def bliss_pdf_report(request):
    registros = Bliss.objects.all()
    template = get_template('bliss/report_pdf.html')
    html = template.render({'registros': registros})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    return response

# Resumo de Situações
def bliss_summary(request):
    registros = Bliss.objects.all()

    resumo_dict = {}
    total_valor_tabela = Decimal('0')
    total_valor_venda = Decimal('0')
    total_unidades = 0

    for reg in registros:
        # valores brutos
        valor_tabela = reg.valor_tabela
        valor_venda = reg.valor_venda
        unidade = reg.unidade.lower()
        situacao = reg.situacao

        # Caso normal (não é loja)
        if unidade != 'loja':
            if situacao not in resumo_dict:
                resumo_dict[situacao] = {
                    'situacao': situacao,
                    'total_valor_tabela': Decimal('0'),
                    'total_valor_venda': Decimal('0'),
                    'total_unidades': 0,
                }

            resumo_dict[situacao]['total_valor_tabela'] += valor_tabela
            resumo_dict[situacao]['total_valor_venda'] += valor_venda
            resumo_dict[situacao]['total_unidades'] += 1

            total_valor_tabela += valor_tabela
            total_valor_venda += valor_venda
            total_unidades += 1

        # Caso especial: unidade == 'Loja'
        else:
            valor_permuta_tabela = valor_tabela * Decimal('0.12826')
            valor_permuta_venda = valor_venda * Decimal('0.12826')

            valor_restante_tabela = valor_tabela * Decimal('0.87174')
            valor_restante_venda = valor_venda * Decimal('0.87174')

            # Grupo "Permuta"
            if 'Permuta' not in resumo_dict:
                resumo_dict['Permuta'] = {
                    'situacao': 'Permuta',
                    'total_valor_tabela': Decimal('0'),
                    'total_valor_venda': Decimal('0'),
                    'total_unidades': 0,
                }

            resumo_dict['Permuta']['total_valor_tabela'] += valor_permuta_tabela
            resumo_dict['Permuta']['total_valor_venda'] += valor_permuta_venda
            resumo_dict['Permuta']['total_unidades'] += 1

            # Grupo restante — se situação == Permuta, o restante vai para "Disponível"
            grupo_restante = 'Disponível' if situacao.lower() == 'permuta' else situacao

            if grupo_restante not in resumo_dict:
                resumo_dict[grupo_restante] = {
                    'situacao': grupo_restante,
                    'total_valor_tabela': Decimal('0'),
                    'total_valor_venda': Decimal('0'),
                    'total_unidades': 0,
                }

            resumo_dict[grupo_restante]['total_valor_tabela'] += valor_restante_tabela
            resumo_dict[grupo_restante]['total_valor_venda'] += valor_restante_venda
            resumo_dict[grupo_restante]['total_unidades'] += 1

            # Totais gerais
            total_valor_tabela += valor_tabela
            total_valor_venda += valor_venda
            total_unidades += 1

    # Percentuais
    resumo = []
    for item in resumo_dict.values():
        item['pct_valor_tabela'] = (item['total_valor_tabela'] / total_valor_tabela * 100) if total_valor_tabela else 0
        item['pct_valor_venda'] = (item['total_valor_venda'] / total_valor_venda * 100) if total_valor_venda else 0
        item['pct_unidades'] = (item['total_unidades'] / total_unidades * 100) if total_unidades else 0
        resumo.append(item)

    # Resumo das lojas
   # Lojas
    lojas = registros.filter(unidade__iexact='loja')
    qtd_lojas = lojas.count()
    m2_lojas = (lojas.aggregate(area=models.Sum('area_privativa'))['area'] or Decimal('0')) * Decimal('0.87174')

    venda_lojas = Decimal('0')

    for loja in lojas:
        fator = Decimal('0.12826') if loja.situacao.lower() == 'permuta' else Decimal('0.87174')
        venda_lojas += loja.valor_tabela * fator

    valor_m2_lojas = venda_lojas / m2_lojas if m2_lojas else Decimal('0')

    # Tipos (exceto loja)
    tipos = registros.filter(situacao__iexact='disponível').exclude(unidade__iexact='loja')
    qtd_tipos = tipos.count()
    m2_tipos = tipos.aggregate(area=models.Sum('area_privativa'))['area'] or Decimal('0')
    venda_tipos = tipos.aggregate(total=models.Sum('valor_tabela'))['total'] or Decimal('0')
    valor_m2_tipos = venda_tipos / m2_tipos if m2_tipos else Decimal('0')
    preco_medio_tipo = venda_tipos / qtd_tipos if qtd_tipos else Decimal('0')

    # Totais
    qtd_total = qtd_lojas + qtd_tipos
    m2_total = m2_lojas + m2_tipos
    venda_total = venda_lojas + venda_tipos
    valor_m2_total = venda_total / m2_total if m2_total else Decimal('0')

    resumo_lojas = [
        {
            'preco_medio_tipo': None,
            'quantidade': qtd_lojas,
            'tipo': 'Loja',
            'm2': m2_lojas,
            'valor_venda': venda_lojas,
            'valor_m2': valor_m2_lojas
        },
        {
            'preco_medio_tipo': preco_medio_tipo,
            'quantidade': qtd_tipos,
            'tipo': 'Tipos',
            'm2': m2_tipos,
            'valor_venda': venda_tipos,
            'valor_m2': valor_m2_tipos
        },
        {
            'preco_medio_tipo': None,
            'quantidade': '',
            'tipo': 'Total',
            'm2': m2_total,
            'valor_venda': venda_total,
            'valor_m2': valor_m2_total
        }
    ]


    return render(request, 'bliss/summary.html', {
        'resumo': resumo,
        'totais': {
            'valor_tabela': total_valor_tabela,
            'valor_venda': total_valor_venda,
            'unidades': total_unidades,
        },
        'resumo_lojas': resumo_lojas,
    })

def atualizar_situacoes(request):
    updates = {
        'QA': [
            ('1-SUN', '201-SUN'),
            ('1-SUN', '206-SUN'),
            ('2-SHINE', '501-SHINE'),
        ],
        'Bloqueada': [
            ('1-SUN', '306-SUN'),
            ('1-SUN', '406-SUN'),
            ('2-SHINE', '305-SHINE'),
            ('2-SHINE', '405-SHINE'),
        ],
        'Permuta': [
            ('1-SUN', '101-SUN'),
            ('1-SUN', '303-SUN'),
            ('1-SUN', '505-SUN'),
            ('1-SUN', '701-SUN'),
            ('1-SUN', '802-SUN'),
            ('2-SHINE', '102-SHINE'),
            ('2-SHINE', '302-SHINE'),
            ('2-SHINE', '401-SHINE'),
            ('2-SHINE', '703-SHINE'),
        ],
        'Permuta/Venda': [
            ('1-SUN', 'Loja'),
        ]
    }

    total_atualizados = 0

    for situacao, unidades in updates.items():
        for bloco, unidade in unidades:
            obj = Bliss.objects.filter(bloco=bloco, unidade=unidade).first()
            if obj:
                obj.situacao = situacao
                obj.save()
                total_atualizados += 1

    messages.success(request, f'{total_atualizados} registros atualizados com sucesso.')
    return redirect('index')